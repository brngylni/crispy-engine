from PIL import Image as img
import PyQt5
from PyQt5 import QtWidgets
import sys
import openpyxl
from PyQt5.QtCore import QPointF, Qt
from PyQt5.uic.properties import QtGui
from openpyxl.styles import Font
from Barcode import Barcode
from ProductsDal import ProductsDal
from PyQt5.QtGui import QPixmap, QImage, QPainter
from PyQt5.QtWidgets import QInputDialog, QLineEdit, QFileDialog, QMessageBox, QTableWidgetItem, QGraphicsPixmapItem
from Analysis import Analysis
from mainWindow import Ui_MainWindow
from MailDal import MailDal
from GroupsDal import GroupsDal
from ProvidersDal import ProvidersDal
import random
from Print import Print

class App(QtWidgets.QMainWindow):

    def __init__(self):
        super(App, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.graph()
        self.queue = ""
        self.ui.inputDialog()
        self.printer = Print()
        self.printers = self.printer.printers
        self.buttons()
        self.initMailList()
        self.initGroup()
        self.initProviders()
        self.initSenders()
        try:
            self.mail = ProvidersDal.getSenders()[0][1]
        except Exception:
            pass
        self.image = [""] * 2
        self.image[0] = [""] * 2
        self.image[1] = [""] * 2
        self.initAnalysis()
        self.initCatalogue()
        self.barcodes = []


    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Enter or event.key() == Qt.Key_Return:
            if self.ui.stackedWidget.currentIndex() == 1:
                self.ui.btnSearchInward.click()
            elif self.ui.stackedWidget.currentIndex() == 2:
                self.ui.btnSearchOutward.click()
            elif self.ui.stackedWidget.currentIndex() == 5:
                self.ui.btnSearchEdit.click()

    def addMail(self):
        mail, okPressed = QInputDialog.getText(self, "Add Mail", "Enter a Mail to Add", QLineEdit.Normal, "")

        if okPressed and mail != "":
            MailDal.addMail(mail)
            self.ui.mailList.addItem(mail)

    def removeMail(self, row):
        MailDal.removeMail(self.ui.mailList.item(row).text())
        self.ui.mailList.takeItem(row)

    def initMailList(self):
        mails = MailDal.getMails()
        self.ui.mailList.clear()


        try:
            for mail in mails:
                self.ui.mailList.addItem(mail[1])
        except Exception:
            self.ui.mailList.addItem(mail[1])


    def initGroup(self):
        self.ui.groupsList.clear()
        companies = GroupsDal.getGroups()
        try:
            for company in companies:
                self.ui.groupsList.addItem(company[1])
        except:
            self.ui.groupsList.addItem(companies[1])

    def addCompany(self):

        name, okPressed = QInputDialog.getText(self, "Add Company", "Enter Company Name", QLineEdit.Normal, "")
        if okPressed and name != "":
            try:
                ProductsDal.addColumn(name)
                GroupsDal.addCompanie(name)
                self.ui.groupsList.addItem(name)
            except:
                QMessageBox.warning(self, "Error", "Already exists.")

    def removeCompany(self, row):
        GroupsDal.removeCompany(self.ui.groupsList.item(row).text())
        ProductsDal.removeColumn(self.ui.groupsList.item(row).text())
        self.ui.groupsList.takeItem(row)

    def buttons(self):
        self.ui.btnHome.clicked.connect(self.switch)
        self.ui.btnInward.clicked.connect(self.switch)
        self.ui.btnOutward.clicked.connect(self.switch)
        self.ui.btnAnalysis.clicked.connect(self.switch)
        self.ui.btnNewProduct.clicked.connect(self.switch)
        self.ui.btnEditProduct.clicked.connect(self.switch)
        self.ui.btnCatalogue.clicked.connect(self.switch)
        self.ui.btnGroup.clicked.connect(self.switch)
        self.ui.btnPrint.clicked.connect(self.switch)
        self.ui.btnAddMailMain.clicked.connect(self.addMail)
        self.ui.btnRemoveMailMain.clicked.connect(lambda : self.removeMail(self.ui.mailList.currentRow()))
        self.ui.btnAddCompany.clicked.connect(self.addCompany)
        self.ui.btnRemoveCompany.clicked.connect(lambda: self.removeCompany(self.ui.groupsList.currentRow()))
        self.ui.btnMailConfig.clicked.connect(self.addSender)
        self.ui.btnSelectSender.clicked.connect(self.selectSender)
        self.ui.btnImage1.clicked.connect(self.image)
        self.ui.btnImage2.clicked.connect(self.image2)
        self.ui.btnImage1Edit.clicked.connect(self.image)
        self.ui.btnImage2Edit.clicked.connect(self.image2)
        self.ui.btnAddProduct.clicked.connect(self.newProduct)
        self.ui.btnRemoveSender.clicked.connect(self.removeSender)
        self.ui.btnSearchInward.clicked.connect(self.searchInward)
        self.ui.btnUpdateInward.clicked.connect(self.updateInward)
        self.ui.btnSearchOutward.clicked.connect(self.searchOutward)
        self.ui.btnOutwardPush.clicked.connect(self.outwardPush)
        self.ui.btnSearchEdit.clicked.connect(self.searchEdit)
        self.ui.btnUpdateEdit.clicked.connect(self.updateEdit)
        self.ui.btnExportInward.clicked.connect(self.exportInward)
        self.ui.btnExportOutward.clicked.connect(self.exportOutward)
        self.ui.btnNextCatalogue.clicked.connect(self.nextCatalogue)
        self.ui.btnPreviousCatalogue.clicked.connect(self.previousCatalogue)
        self.ui.productListCatalogue.itemSelectionChanged.connect(self.catalogue)
        self.ui.pushButton_2.clicked.connect(self.generateBarcode)
        self.ui.ui.btnPrint.clicked.connect(self.print)
        self.ui.inputUi.btnOk.clicked.connect(self.inputDialog)
        self.ui.btnExportCatalogue.clicked.connect(self.exportCatalogue)

    def switch(self):
        sender = self.sender()
        if sender.text() == "Home":
            self.ui.stackedWidget.setCurrentWidget(self.ui.pageHome)
            self.initMailList()
            self.initProviders()
            self.initSenders()
        elif sender.text() == "Inward":
            self.ui.stackedWidget.setCurrentWidget(self.ui.pageInward)
        elif sender.text() == "Outward":
            self.ui.stackedWidget.setCurrentWidget(self.ui.pageOutward)
        elif sender.text() == "Analysis":
            self.ui.stackedWidget.setCurrentWidget(self.ui.pageAnalysis)
            self.initAnalysis()
        elif sender.text() == "New Product":
            self.ui.stackedWidget.setCurrentWidget(self.ui.pageNewProduct)
        elif sender.text() == "Edit Product":
            self.ui.stackedWidget.setCurrentWidget(self.ui.pageEditProduct)
        elif sender.text() == "Catalogue":
            self.ui.stackedWidget.setCurrentWidget(self.ui.pageCatalogue)
            self.initCatalogue()
        elif sender.text() == "Group":
            self.ui.stackedWidget.setCurrentWidget(self.ui.pageGroup)
            self.initGroup()
        elif sender.text() == "Print":
            self.ui.stackedWidget.setCurrentWidget(self.ui.page)


    def initProviders(self):
        self.ui.providersMailConfig.addItem("Google")
        self.ui.providersMailConfig.addItem("Outlook")
        self.ui.providersMailConfig.addItem("Yahoo.")

    def addSender(self):
        mail = self.ui.mailConfigAddress.text()
        password = self.ui.mailConfigPassword.text()
        provider = self.ui.providersMailConfig.currentText()
        ProvidersDal.addSender(mail, password, provider)
        self.ui.senderList.addItem(mail)

    def initSenders(self):
        senders = ProvidersDal.getSenders()
        self.ui.senderList.clear()
        for sender in senders:
            self.ui.senderList.addItem(sender[1])

    def selectSender(self):
        try:
            address = self.ui.senderList.currentItem().text()
            self.mail = address
        except Exception:
            pass

    def image(self):

        self.image[0] = QFileDialog.getOpenFileName(self, 'Choose Image', "/")
        self.ui.imageLabelNewProduct.setPixmap(QPixmap(self.image[0][0]))
        if self.image[0] is None or self.image[0] == '':
            self.image[0] = [""] * 2
    def image2(self):
        self.image[1] = QFileDialog.getOpenFileName(self, "Choose Image", "/")
        self.ui.imageLabel2NewProduct.setPixmap(QPixmap(self.image[1][0]))
        if self.image[1] is None or self.image[1] == '':
            self.image[1] = [""] * 2
    def newProduct(self):
            barcode = self.ui.barcodeNumberFieldNew.text()
            name = self.ui.productNameFieldNew.text()
            quantity = int(self.ui.quantityFieldNew.text())
            if barcode == "" or name == "":
                QMessageBox.warning(self, "Error", "Please fill the blanks.")
            else:
                print(self.image)
                if  ProductsDal.newProduct(barcode, name, quantity,self.image[0][0],self.image[1][0], self.ui.inputs, self.ui.labels):
                    QMessageBox.information(self, "Succesful", "Product Added Succesfully.")
                    self.ui.imageLabelNewProduct.clear()
                    self.ui.imageLabel2NewProduct.clear()
                    self.ui.barcodeNumberFieldNew.clear()
                    self.ui.quantityFieldNew.clear()
                else:
                    QMessageBox.warning(self, "Error", "Please fill the blanks.")
            self.image[0] = [""] * 2
            self.image[1] = [""] * 2

    def removeSender(self):
        mail = self.ui.senderList.currentItem()
        ProvidersDal.removeSender(mail.text())
        self.ui.senderList.takeItem(self.ui.senderList.currentIndex().row())

    def searchInward(self):
        barcode = self.ui.barcodeNumbeFieldInward.text()
        product = ProductsDal.searchProduct(barcode)
        if product is not None:
            try:
                self.ui.productNameFieldInward.setText(product[1])
                self.ui.currentQuantityFieldInward.setText(str(product[2]))
                self.ui.imageLabelInward.setPixmap(QPixmap(product[7]))
                self.ui.imageLabel2Inward.setPixmap(QPixmap(product[8]))
                i=10
                while i<len(product):
                    self.ui.inwardSkusField[i-10].setText(product[i])
                    i += 1

            except Exception:pass
        else:
            self.ui.productNameFieldInward.clear()
            self.ui.currentQuantityFieldInward.clear()
            for i in self.ui.inwardSkusField:
                i.clear()
            self.ui.imageLabel2Inward.clear()
            self.ui.imageLabelInward.clear()

    def updateInward(self):
        barcodeNumber = self.ui.barcodeNumbeFieldInward.text()
        amount = int(self.ui.inwardAmountFieldInward.text())
        finalAmount = ProductsDal.updateInward(self, amount, barcodeNumber)
        QMessageBox.information(self, "Success", f"Updated Succesfuly\nCurrent Stock =  {finalAmount}")
        self.ui.currentQuantityFieldInward.clear()
        self.ui.barcodeNumbeFieldInward.clear()
        self.ui.productNameFieldInward.clear()
        self.ui.inwardAmountFieldInward.clear()
        self.ui.imageLabelInward.clear()
        self.ui.imageLabel2Inward.clear()

    def searchOutward(self):
        barcode = self.ui.barcodeNumberFieldOutward.text()
        product = ProductsDal.searchProduct(barcode)
        if product is not None:
            try:
                self.ui.productNameFieldOutward.setText(product[1])
                self.ui.quantityFieldOutward.setText(str(product[2]))
                self.ui.inwardDateFieldOutward.setText(product[3])
                self.ui.imageLabelOutward.setPixmap(QPixmap(product[7]))
                self.ui.imageLabel2Outward.setPixmap(QPixmap(product[8]))
                i = 10
                while i < len(product):
                    self.ui.outwardFields[i - 10].setText(product[i])
                    i += 1

            except Exception:
                pass
        else:
            self.ui.productNameFieldOutward.clear()
            self.ui.quantityFieldOutward.clear()
            self.ui.inwardDateFieldOutward.clear()
            for i in self.ui.outwardFields:
               i.clear()
            self.ui.imageLabel2Outward.clear()
            self.ui.imageLabelOutward.clear()

    def outwardPush(self):
        barcode = str(self.ui.barcodeNumberFieldOutward.text())
        amount = int(self.ui.outwardAmountFieldOutward.text())
        if ProductsDal.outward(self, barcode, amount):
            finalAmount = ProductsDal.searchProduct(barcode)[2]
            QMessageBox.information(self, "Success", f"Outward Succesful\nCurrent Amount = {str(finalAmount)} ")
            self.ui.quantityFieldOutward.clear()
            self.ui.barcodeNumberFieldOutward.clear()
            self.ui.outwardAmountFieldOutward.clear()
            self.ui.productNameFieldOutward.clear()
            self.ui.inwardDateFieldOutward.clear()
            self.ui.imageLabelOutward.clear()
            self.ui.imageLabel2Outward.clear()
        else:
            QMessageBox.warning(self, "Error", "Error")

    def initAnalysis(self):
        Analysis.analyse()
        self.ui.chart.setPixmap(QPixmap("saleRates.png"))

    def searchEdit(self):
        barcode = self.ui.barcodeNumberFieldEdit.text()
        product = ProductsDal.searchProduct(barcode)
        if product is not None:
            try:
                self.ui.productNameFieldEdit.setText(product[1])
                self.ui.quantityFieldEdit.setText(str(product[2]))
                self.ui.imageLabelEdit.setPixmap(QPixmap(product[7]))
                self.ui.imageLabel2Edit.setPixmap(QPixmap(product[8]))
                self.ui.inwardDateFieldEdit.setText(product[3])
                self.ui.soldFieldEdit.setValue(product[4])

                self.image[0][0] = product[7]
                self.image[1][0] = product[8]
                i = 10
                while i < len(product):
                    self.ui.skusEditField[i - 10].setText(product[i])
                    i += 1

            except:
                pass
        else:
            self.ui.productNameFieldEdit.clear()
            self.ui.quantityFieldEdit.clear()
            self.ui.inwardDateFieldEdit.clear()
            self.ui.soldFieldEdit.clear()
            self.ui.imageLabelEdit.clear()
            self.ui.imageLabel2Edit.clear()
            for i in self.ui.skusEditField:
                i.clear()

    def updateEdit(self):
        barcode = self.ui.barcodeNumberFieldEdit.text()
        product = ProductsDal.searchProduct(barcode)
        name = self.ui.productNameFieldEdit.text()
        dateInward = self.ui.inwardDateFieldEdit.text()
        sold = self.ui.soldFieldEdit.text()
        if product is not None:
            ProductsDal.edit(barcode, name, dateInward, sold, self.image[0][0], self.image[1][0], self.ui.skusEditLabel, self.ui.skusEditField)
            QMessageBox.information(self, "Success", "Edited Succesfully")
            self.ui.imageLabelEdit.clear()
            self.ui.imageLabel2Edit.clear()
            self.ui.barcodeNumberFieldEdit.clear()
            self.ui.productNameFieldEdit.clear()
            self.ui.quantityFieldEdit.clear()
            self.ui.inwardDateFieldEdit.clear()
            self.ui.soldFieldEdit.clear()

    def exportCatalogue(self):
        path = QFileDialog.getExistingDirectory(self, "Select a location")
        products = ProductsDal.getProducts()
        groups = GroupsDal.getGroups()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Inward Data"
        a1 = sheet["A1"]
        a1.font = Font(bold=True)
        a1.value = "Barcode Number"
        b1 = sheet["B1"]
        b1.font = Font(bold=True)
        b1.value = "Name"
        c1 = sheet["C1"]
        c1.font = Font(bold=True)
        c1.value = "Quantity"
        i = 4
        for group in groups:
            sheet.cell(row=1, column=i).font = Font(bold=True)
            sheet.cell(row=1, column=i, value=group[1])
            i += 1
        j = 2
        k = 4
        for product in products:
            sheet.cell(row=j, column=1, value=product[6])
            sheet.cell(row=j, column=2, value=product[1])
            sheet.cell(row=j, column=3, value=product[2])
            while k < i:
                sheet.cell(row=j, column=k, value=product[k + 6])
                k += 1
            k = 4
            j += 1
        workbook.save(f"{path}/productExport.xlsx")
        workbook.close()

    def exportInward(self):
        path = QFileDialog.getExistingDirectory(self, "Select a location")
        products = ProductsDal.getProducts()
        groups = GroupsDal.getGroups()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Inward Data"
        a1 = sheet["A1"]
        a1.font = Font(bold=True)
        a1.value = "Barcode Number"
        b1 = sheet["B1"]
        b1.font = Font(bold=True)
        b1.value = "Inward Date"
        c1 = sheet["C1"]
        c1.font = Font(bold=True)
        c1.value = "Quantity"
        i=4
        for group in groups:
            sheet.cell(row=1, column=i).font =Font(bold=True)
            sheet.cell(row=1, column=i, value=group[1])
            i += 1
        j = 2
        k = 4
        for product in products:
            sheet.cell(row=j, column=1, value=product[6])
            sheet.cell(row=j, column=2, value=product[3])
            sheet.cell(row=j, column=3, value=product[2])
            while k < i:
                sheet.cell(row=j, column=k, value=product[k+6])
                k += 1
            k = 4
            j += 1
        workbook.save(f"{path}/inwardExport.xlsx")
        workbook.close()

    def exportOutward(self):

        path = QFileDialog.getExistingDirectory(self, "Select a location")
        products = ProductsDal.getProducts()
        groups = GroupsDal.getGroups()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Outward Data"
        a1 = sheet["A1"]
        a1.font = Font(bold=True)
        a1.value = "Barcode Number"
        b1 = sheet["B1"]
        b1.font = Font(bold=True)
        b1.value = "Outward Date"
        c1 = sheet["C1"]
        c1.font = Font(bold=True)
        c1.value = "Quantity"
        i = 4
        for group in groups:
            sheet.cell(row=1, column=i).font = Font(bold=True)
            sheet.cell(row=1, column=i, value=group[1])
            i += 1
        j = 2
        k = 4
        for product in products:
            sheet.cell(row=j, column=1, value=product[6])
            sheet.cell(row=j, column=2, value=product[9])
            sheet.cell(row=j, column=3, value=product[2])
            while k < i:
                sheet.cell(row=j, column=k, value=product[k + 6])
                k += 1
            k = 4
            j += 1
        workbook.save(f"{path}/outwardExport.xlsx")
        workbook.close()

    def initCatalogue(self):
        products = ProductsDal.getProducts()
        i=0
        self.ui.productListCatalogue.setRowCount(len(products))
        for product in products:
            barcodeNumber = QTableWidgetItem(str(product[6]))
            name = QTableWidgetItem(str(product[1]))
            quantity = QTableWidgetItem(str(product[2]))
            self.ui.productListCatalogue.setItem(i,0,barcodeNumber)
            self.ui.productListCatalogue.setItem(i,1,name)
            self.ui.productListCatalogue.setItem(i,2,quantity)
            i+= 1

    def catalogue(self):
        index = self.ui.productListCatalogue.currentRow()
        barcode = self.ui.productListCatalogue.item(index, 0).text()
        product = ProductsDal.searchProduct(barcode)

        try:
            image1 = QPixmap(product[7])
            image2 = QPixmap(product[8])
            self.ui.imageLabel1Catalogue.setPixmap(image1)
            self.ui.imageLabel2Catalogue.setPixmap(image2)
        except Exception:
            self.ui.imageLabel1Catalogue.clear()
            self.ui.imageLabel2Catalogue.clear()

    def nextCatalogue(self):
        row = self.ui.productListCatalogue.currentRow()
        col = self.ui.productListCatalogue.currentColumn()
        if row < self.ui.productListCatalogue.rowCount()-1:
            self.ui.productListCatalogue.setCurrentCell(row+1, col)
    def previousCatalogue(self):
        row = self.ui.productListCatalogue.currentRow()
        col = self.ui.productListCatalogue.currentColumn()
        if row >0:
            self.ui.productListCatalogue.setCurrentCell(row -1, col)


    def stacked(self):
        index = self.ui.stackedWidget.currentIndex()

    def generateBarcode(self):
        try:
            barcodeNumber = self.ui.lineEdit_3.text()
            barcodeLocation = Barcode.barcode(barcodeNumber)
            pos = random.randint(0,1000)
            if self.ui.window.isVisible():
                self.ui.ui.graphicsView.scene().addItem(Image(pos, 50, barcodeLocation))
            else:
                self.ui.window.show()
                self.ui.ui.graphicsView.scene().addItem(Image(pos, 50, barcodeLocation))
        except Exception:
            QMessageBox.warning(self, "Error", "Please enter a barcode number.")

    def print(self):
        for printer in self.printers:
            self.ui.inputUi.printersList.addItem(printer)

        image = QImage(1600, 600, QImage.Format_Mono)
        painter = QPainter(image)
        self.ui.ui.graphicsView.scene().render(painter)
        painter.end()
        image.save("barcodes/toPrinter.png")
        toPcx = img.open("barcodes/toPrinter.png")
        toPcx.save("barcodes/toPrinter.pcx")

        self.ui.inputWindow.show()

    def inputDialog(self):
        try:
            self.labelWidth = 203 * float(self.ui.inputUi.widthField.text())
            self.labelHeight = 203 * float(self.ui.inputUi.heightField.text())
            self.labelGap = 203 * float(self.ui.inputUi.gapField.text())
            self.queue = self.ui.inputUi.printersList.currentText()
            self.ui.inputWindow.close()
            self.printer.print(self.queue, (self.labelHeight, self.labelGap), self.labelWidth)
        except Exception:
           QMessageBox.warning(self, "Error", "İnvalid Parameter.")








class Image(QGraphicsPixmapItem):
     def __init__(self, x, y, pixmap):
         super(Image, self).__init__()
         self.setPixmap(QPixmap(pixmap))
         self.setPos(x, y)
         self.setAcceptHoverEvents(True)

     def hoverEnterEvent(self, event: 'QGraphicsSceneHoverEvent') -> None:
         app.instance().setOverrideCursor(Qt.OpenHandCursor)

     def hoverLeaveEvent(self, event: 'QGraphicsSceneHoverEvent') -> None:
         app.instance().restoreOverrideCursor()

     def mousePressEvent(self, event: 'QGraphicsSceneMouseEvent') -> None:
         pass

     def mouseMoveEvent(self, event: 'QGraphicsSceneMouseEvent') -> None:
         orig_cursor_position = event.lastScenePos()
         updated_cursor_position = event.scenePos()
         orig_position = self.scenePos()

         updated_cursor_x = updated_cursor_position.x() - orig_cursor_position.x() + orig_position.x()
         updated_cursor_y = updated_cursor_position.y() - orig_cursor_position.y() + orig_position.y()

         self.setPos(QPointF(updated_cursor_x, updated_cursor_y))

     def mouseReleaseEvent(self, event):
        pass






app = QtWidgets.QApplication(sys.argv)
win = App()
win.show()
sys.exit(app.exec_())