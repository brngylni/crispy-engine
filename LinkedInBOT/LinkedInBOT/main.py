# Importing necessary modules.
from LinkedIn import LinkedIn
import openpyxl
import time


# Opening the config file for necessary information.
config = open("config.txt", "r")
lines = config.readlines()
username = lines[0].split()[2]
password = lines[1].split()[2]
runtime = lines[2].split()[2]
config.close()
# Opening the excel file.
book = openpyxl.load_workbook("workbook.xlsx")
sheet = book.active
# Creating Linkedin Object.
linkedin = LinkedIn(username, password)


def initialization():
    # Sign in to linkedin with given information in which config.txt.
    linkedin.login()
# Fetching the connections according to runtime counter in which config.txt.
    connections = linkedin.fetchConnections(runtime)
    row = 1
# Controlling the connections for accepted connections and updating the excel sheet.
    while row <= int(runtime)*90:
        if sheet[f"H{row}"].value in connections:
            sheet[f"J{row}"].value = u'\u2713'
            if sheet[f"K{row}"].value is None:
                url = sheet[f"H{row}"].value
                linkedin.sendFollowUpMessage(url)
                sheet[f"K{row}"].value = u'\u2713'
        row += 1


def finalization():
    # Final operations. Closing browser, closing files and rewriting configs.
    book.save("workbook.xlsx")
    linkedin.closeBrowser()
    new_runtime = int(runtime) + 1
    file = open("config.txt", "w")
    file.write(f"Username = {username}\nPassword = {password}\nRuntime = {new_runtime}\n")
    file.close()


# Calling initialization function.
initialization()
# Calculating the number of rows that processed before this run.
processed = int(runtime)*90
# Handling the row exception.
if processed == 0:
    for i in range(1, 90):
        time.sleep(1)
        if sheet[f"H{i}"].value is not None:
            if sheet[f"I{i}"].value is None:
                url = sheet[f"H{i}"].value
                linkedin.sendConnectionRequest(url)
                sheet[f"I{i}"].value = u'\u2713'



else:
    if ((processed + 90) <= (sheet.max_row)):
        for i in range(processed, processed+90):
            time.sleep(1)
            url = sheet[f"H{i}"].value
            linkedin.sendConnectionRequest(url)
            sheet[f"I{i}"].value = u'\u2713'
    else:
        for i in range(processed, sheet.max_row):
            time.sleep(1)
            url = sheet[f"H{i}"].value
            linkedin.sendConnectionRequest(url)
            sheet[f"I{i}"].value = u'\u2713'

finalization()
